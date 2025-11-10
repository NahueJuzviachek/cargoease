# home/views.py
from datetime import datetime, timedelta, date, time as dtime
from calendar import month_name, monthrange
from collections import defaultdict
from django.utils.dateparse import parse_date
from django.http import HttpResponse
import calendar
import csv
from django.db.models import (
    Sum, F, Value, CharField, Case, When, Count,
    DecimalField, ExpressionWrapper
)
from django.db.models.functions import Coalesce, Concat, TruncMonth
from django.http import JsonResponse
from django.shortcuts import render
from django.utils import timezone
from django.views.decorators.http import require_GET
from io import BytesIO
from viajes.models import Viaje
from vehiculos.models import Vehiculo
from aceite.models import Aceite
from neumaticos.models import Neumatico

# Importa Cliente solo si existe
try:
    from clientes.models import Cliente
except Exception:
    Cliente = None

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
except Exception:
    # reportlab puede no estar instalado
    SimpleDocTemplate = None


# =========================
# Vista principal de Home
# =========================
def home(request):
    """Renderiza la página principal del dashboard"""
    return render(request, "home/home.html")


# =========================
# Helpers internos
# =========================
def _parse_date(s: str):
    """Convierte un string 'YYYY-MM-DD' a date, retorna None si falla"""
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None


def _filtrar_por_fecha_viaje(qs, desde_date: date, hasta_date: date):
    """
    Filtra un queryset de Viaje por rango de fechas.
    Soporta distintos esquemas de campo fecha (aware, naive, datetime, date).
    """
    if not hasattr(Viaje, "fecha"):
        return qs
    try:
        # Intento filtro simple por date
        return qs.filter(fecha__date__range=(desde_date, hasta_date))
    except Exception:
        pass
    try:
        # Intento filtro con aware datetime
        tz = timezone.get_current_timezone()
        desde_dt = timezone.make_aware(datetime.combine(desde_date, dtime.min), tz)
        hasta_dt = timezone.make_aware(datetime.combine(hasta_date, dtime.max), tz)
        return qs.filter(fecha__range=(desde_dt, hasta_dt))
    except Exception:
        return qs


def _iter_months(y1: int, m1: int, y2: int, m2: int):
    """Genera (año, mes) entre dos fechas mensuales inclusivas"""
    y, m = y1, m1
    while (y < y2) or (y == y2 and m <= m2):
        yield (y, m)
        m += 1
        if m > 12:
            m = 1
            y += 1


def _neumaticos_vigentes_qs():
    """
    Devuelve un queryset de Neumático excluyendo bajas o eliminados.
    Funciona aunque el modelo tenga distintos campos para estado.
    """
    qs = Neumatico.objects.all()
    try:
        qs = qs.filter(activo=True)
    except Exception:
        pass
    try:
        qs = qs.filter(eliminado=False)
    except Exception:
        pass
    try:
        qs = qs.filter(fecha_baja__isnull=True)
    except Exception:
        pass
    return qs


# =========================
# JSON: Clientes top mensual
# =========================
@require_GET
def data_clientes_top_mensual(request):
    """
    Devuelve JSON con los 3 clientes con más viajes por mes en un rango determinado.
    Incluye labels de meses y series de conteo de viajes.
    """
    hoy = timezone.localdate()
    meses_default = int(request.GET.get("meses", "6"))

    def _parse_month(s):
        """Convierte 'YYYY-MM' a (año, mes)"""
        try:
            y, m = s.split("-")
            return int(y), int(m)
        except Exception:
            return None

    desde_param = request.GET.get("desde")
    hasta_param = request.GET.get("hasta")

    if desde_param and hasta_param and _parse_month(desde_param) and _parse_month(hasta_param):
        y1, m1 = _parse_month(desde_param)
        y2, m2 = _parse_month(hasta_param)
    else:
        # Rango por defecto (últimos 'meses_default' meses)
        y2, m2 = hoy.year, hoy.month
        y1, m1 = y2, m2 - (meses_default - 1)
        while m1 <= 0:
            m1 += 12
            y1 -= 1

    def _labels_from_range(y1i, m1i, y2i, m2i):
        """Genera labels de meses tipo 'Enero 2025'"""
        return [f"{month_name[m]} {y}" for (y, m) in _iter_months(y1i, m1i, y2i, m2i)]

    def _run_period(y1i, m1i, y2i, m2i):
        """Calcula top clientes y series de viajes por mes"""
        desde_date = date(y1i, m1i, 1)
        last_day = monthrange(y2i, m2i)[1]
        hasta_date = date(y2i, m2i, last_day)

        qs = Viaje.objects.filter(cliente_id__isnull=False)
        qs = _filtrar_por_fecha_viaje(qs, desde_date, hasta_date)

        # Top 3 clientes por cantidad de viajes
        top_totales = (qs.values('cliente_id')
                         .annotate(total=Count('id'))
                         .order_by('-total')[:3])
        top_ids = [row['cliente_id'] for row in top_totales]
        if not top_ids:
            return _labels_from_range(y1i, m1i, y2i, m2i), []

        # Conteo mensual de viajes para top clientes
        mensuales = (qs.filter(cliente_id__in=top_ids)
                       .annotate(ym=TruncMonth('fecha'))
                       .values('cliente_id', 'ym')
                       .annotate(cnt=Count('id'))
                       .order_by())

        months_list = list(_iter_months(y1i, m1i, y2i, m2i))
        months_index = {(y, m): idx for idx, (y, m) in enumerate(months_list)}
        series_dict = {cid: [0] * len(months_list) for cid in top_ids}

        for row in mensuales:
            ym = row['ym']
            if ym is None:
                continue
            key = (ym.year, ym.month)
            idx = months_index.get(key)
            if idx is not None:
                series_dict[row['cliente_id']][idx] = row['cnt']

        # Mapeo de labels con nombres de clientes
        labels_map = {}
        if Cliente is not None:
            try:
                clientes = Cliente.objects.filter(pk__in=top_ids)
                for c in clientes:
                    label = getattr(c, "nombre", None) or getattr(c, "razon_social", None) or str(c)
                    labels_map[c.pk] = label
            except Exception:
                pass

        labels_final = _labels_from_range(y1i, m1i, y2i, m2i)
        series = [{"label": labels_map.get(cid) or f"Cliente {cid}", "data": series_dict[cid]} for cid in top_ids]

        return labels_final, series

    # Ejecuta cálculo del período principal
    labels, series = _run_period(y1, m1, y2, m2)

    # Si no hay datos, intenta con últimos 12 meses
    if not series or all(sum(s["data"]) == 0 for s in series):
        y2b, m2b = hoy.year, hoy.month
        y1b, m1b = y2b, m2b - (12 - 1)
        while m1b <= 0:
            m1b += 12
            y1b -= 1
        labels, series = _run_period(y1b, m1b, y2b, m2b)

    return JsonResponse({"labels": labels, "series": series})


# =========================
# JSON: Ranking de km por vehículo
# =========================
@require_GET
def data_ranking_km(request):
    """Devuelve JSON con los 3 vehículos que recorrieron más km en un rango de fechas"""
    hoy = timezone.localdate()

    def _sum_distancia(qs):
        """Agrega el total de km recorridos por vehículo"""
        total = Coalesce(
            Sum('distancia'),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=2))
        )
        return (qs.values('vehiculo__id', 'vehiculo__dominio')
                  .annotate(total_km=total)
                  .order_by('-total_km')[:3])

    def _run_range(desde_date, hasta_date):
        """Filtra viajes por rango de fechas y devuelve los top 3 vehículos"""
        qs = Viaje.objects.select_related("vehiculo").all()
        qs = _filtrar_por_fecha_viaje(qs, desde_date, hasta_date)
        agg = _sum_distancia(qs)
        items = []
        for a in agg:
            etiqueta = a.get('vehiculo__dominio') or f"Vehículo {a['vehiculo__id']}"
            items.append({"vehiculo": etiqueta, "km": float(a['total_km'] or 0)})
        return items

    desde_str = request.GET.get("desde")
    hasta_str = request.GET.get("hasta")

    # Rango de fechas por defecto: últimos 6 meses
    if desde_str and hasta_str:
        desde_date = _parse_date(desde_str) or (hoy - timedelta(days=180))
        hasta_date = _parse_date(hasta_str) or hoy
    else:
        desde_date = hoy - timedelta(days=180)
        hasta_date = hoy

    items = _run_range(desde_date, hasta_date)

    # Si no hay datos, intenta con últimos 12 meses
    if not items or all((it["km"] or 0) == 0 for it in items):
        items = _run_range(hoy - timedelta(days=365), hoy)

    return JsonResponse({"items": items})


# =========================
# JSON: Aceite – Top 5 con menos km restantes
# =========================
@require_GET
def data_aceite_top5(request):
    """Devuelve JSON con los 5 aceites con menor km restante antes de cambio"""
    km_restantes_expr = ExpressionWrapper(
        Coalesce(F('vida_util_km'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=2))) -
        Coalesce(F('km_acumulados'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=2))),
        output_field=DecimalField(max_digits=12, decimal_places=2)
    )
    vehiculo_label_expr = Coalesce(
        F('vehiculo__dominio'),
        Concat(Value("Vehículo "), F('vehiculo_id'), output_field=CharField())
    )
    qs = (Aceite.objects
          .annotate(
              km_restantes=km_restantes_expr,
              vehiculo_label=vehiculo_label_expr,
          )
          .order_by('km_restantes')
          .values('vehiculo_label', 'tipo', 'km_restantes')[:5])

    items = [{"vehiculo": a.get('vehiculo_label') or "",
              "tipo": a.get('tipo') or "",
              "km_restantes": float(a.get('km_restantes') or 0)} for a in qs]

    return JsonResponse({"items": items})


# =========================
# JSON: Neumáticos – Estado global de flota
# =========================
@require_GET
def data_neumaticos_estado(request):
    """Devuelve JSON con conteo de neumáticos por estado"""
    agg = (_neumaticos_vigentes_qs()
           .values(nombre=F('estado__descripcion'))
           .annotate(cant=Count('idNeumatico'))
           .order_by('nombre'))

    labels = [a['nombre'] for a in agg]
    values = [a['cant'] for a in agg]
    return JsonResponse({"labels": labels, "values": values})

def reporte_clientes_view(request):
    """
    Genera el reporte de clientes dentro de un rango mensual (YYYY-MM).
    Soporta:
      - HTML (por defecto)
      - CSV  -> ?format=csv
      - Excel (xlsx) -> ?export=excel
      - PDF  -> ?export=pdf

    Parámetros:
      - desde: 'YYYY-MM'
      - hasta: 'YYYY-MM'
    """
    # leer params (GET o POST)
    data = request.GET if request.method == "GET" else request.POST
    desde_m = data.get("desde")
    hasta_m = data.get("hasta")

    # parseo 'YYYY-MM' a fechas
    def month_to_range(ym: str):
        if not ym or "-" not in ym:
            return None, None
        try:
            y, m = ym.split("-")
            y = int(y); m = int(m)
            first_day = date(y, m, 1)
            last_day = date(y, m, calendar.monthrange(y, m)[1])
            return first_day, last_day
        except Exception:
            return None, None

    desde_date, _ = month_to_range(desde_m)
    _, hasta_date = month_to_range(hasta_m)

    # Si no vienen fechas, valor por defecto: últimos 6 meses
    hoy = date.today()
    if not desde_date or not hasta_date:
        y2, m2 = hoy.year, hoy.month
        last_day = calendar.monthrange(y2, m2)[1]
        hasta_date = date(y2, m2, last_day)
        m_back = m2 - 5
        y_back = y2
        while m_back <= 0:
            m_back += 12
            y_back -= 1
        desde_date = date(y_back, m_back, 1)

    # queryset base
    qs = Viaje.objects.filter(cliente_id__isnull=False)
    # filtro fechas usando tu helper si lo tenés; si no, filter directo:
    try:
        # usa tu helper si existe
        qs = _filtrar_por_fecha_viaje(qs, desde_date, hasta_date)
    except Exception:
        # fallback
        qs = qs.filter(fecha__date__range=(desde_date, hasta_date))

    # --- Top clientes por ingresos (sum valor_flete) ---
    clientes_ingresos_qs = (
        qs.values("cliente__id", "cliente__nombre", "cliente__correo")
          .annotate(
              total_ingresos=Coalesce(
                  Sum("valor_flete", output_field=DecimalField(max_digits=18, decimal_places=2)),
                  Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))
              )
          )
          .order_by("-total_ingresos")
    )

    # --- Top clientes por cantidad de viajes ---
    clientes_viajes_qs = (
        qs.values("cliente__id", "cliente__nombre", "cliente__correo")
          .annotate(total_viajes=Count("id"))
          .order_by("-total_viajes")
    )

    # Convertir a listas (útil para export)
    ingresos = list(clientes_ingresos_qs)
    viajes = list(clientes_viajes_qs)

    # Parámetros de export
    formato_csv = data.get("format", "").lower() == "csv"
    export_excel = data.get("export", "").lower() == "excel"
    export_pdf = data.get("export", "").lower() == "pdf"

    # ------------------------
    # CSV (descarga)
    # ------------------------
    if formato_csv:
        filename = f"reporte_clientes_{desde_date.strftime('%Y%m')}_{hasta_date.strftime('%Y%m')}.csv"
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        writer = csv.writer(response)

        writer.writerow([f"Reporte de Clientes: {desde_date} a {hasta_date}"])
        writer.writerow([])
        writer.writerow(["Top por Ingresos"])
        writer.writerow(["cliente_id", "nombre", "correo", "total_ingresos"])
        for r in ingresos:
            writer.writerow([
                r.get("cliente__id"),
                r.get("cliente__nombre") or "",
                r.get("cliente__correo") or "",
                f"{float(r.get('total_ingresos') or 0):.2f}"
            ])

        writer.writerow([])
        writer.writerow(["Top por Cantidad de Viajes"])
        writer.writerow(["cliente_id", "nombre", "correo", "total_viajes"])
        for r in viajes:
            writer.writerow([
                r.get("cliente__id"),
                r.get("cliente__nombre") or "",
                r.get("cliente__correo") or "",
                int(r.get("total_viajes") or 0)
            ])

        return response

    # ------------------------
    # EXCEL (.xlsx)
    # ------------------------
    if export_excel:
        if openpyxl is None:
            return HttpResponse("openpyxl no está instalado. Instalar con `pip install openpyxl`", status=500)

        wb = openpyxl.Workbook()
        # Hoja 1: Ingresos
        ws1 = wb.active
        ws1.title = "Ingresos"
        headers1 = ["cliente_id", "nombre", "correo", "total_ingresos"]
        ws1.append(["Reporte de Clientes", f"{desde_date} → {hasta_date}"])
        ws1.append([])
        ws1.append(headers1)
        for r in ingresos:
            ws1.append([
                r.get("cliente__id"),
                r.get("cliente__nombre") or "",
                r.get("cliente__correo") or "",
                float(r.get("total_ingresos") or 0),
            ])
        # Ajustar ancho columnas
        for i, col in enumerate(headers1, 1):
            ws1.column_dimensions[get_column_letter(i)].width = max(12, len(col) + 2)

        # Hoja 2: Viajes
        ws2 = wb.create_sheet(title="Cantidad de Viajes")
        headers2 = ["cliente_id", "nombre", "correo", "total_viajes"]
        ws2.append(["Top por Cantidad de Viajes"])
        ws2.append([])
        ws2.append(headers2)
        for r in viajes:
            ws2.append([
                r.get("cliente__id"),
                r.get("cliente__nombre") or "",
                r.get("cliente__correo") or "",
                int(r.get("total_viajes") or 0),
            ])
        for i, col in enumerate(headers2, 1):
            ws2.column_dimensions[get_column_letter(i)].width = max(12, len(col) + 2)

        # Guardar en memoria y devolver
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        filename = f"reporte_clientes_{desde_date.strftime('%Y%m')}_{hasta_date.strftime('%Y%m')}.xlsx"
        response = HttpResponse(stream.read(),
                                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    # ------------------------
    # PDF (reportlab)
    # ------------------------
    if export_pdf:
        if SimpleDocTemplate is None:
            return HttpResponse("reportlab no está instalado. Instalar con `pip install reportlab`", status=500)

        buffer = BytesIO()
        # Documento en horizontal para tablas largas
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph(f"Reporte de Clientes: {desde_date} → {hasta_date}", styles["Title"]))
        elements.append(Spacer(1, 12))

        # Tabla 1: Ingresos
        data_ing = [["#", "Cliente", "Correo", "Total Ingresos"]]
        for i, r in enumerate(ingresos, start=1):
            data_ing.append([i, r.get("cliente__nombre") or "", r.get("cliente__correo") or "", f"{float(r.get('total_ingresos') or 0):.2f}"])

        t1 = Table(data_ing, colWidths=[30, 200, 220, 100])
        t1.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2d6a4f")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ]))
        elements.append(t1)
        elements.append(Spacer(1, 18))

        # Tabla 2: Viajes
        data_via = [["#", "Cliente", "Correo", "Total Viajes"]]
        for i, r in enumerate(viajes, start=1):
            data_via.append([i, r.get("cliente__nombre") or "", r.get("cliente__correo") or "", int(r.get("total_viajes") or 0)])

        t2 = Table(data_via, colWidths=[30, 200, 220, 80])
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#114b8b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ]))
        elements.append(t2)

        doc.build(elements)
        buffer.seek(0)
        filename = f"reporte_clientes_{desde_date.strftime('%Y%m')}_{hasta_date.strftime('%Y%m')}.pdf"
        return HttpResponse(buffer.read(), content_type="application/pdf", headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        })

    # ------------------------
    # Si no piden export, render HTML
    # ------------------------
    context = {
        "clientes_ingresos": clientes_ingresos_qs,
        "clientes_viajes": clientes_viajes_qs,
        "desde": desde_date,
        "hasta": hasta_date,
    }
    return render(request, "home/reporte_clientes.html", context)